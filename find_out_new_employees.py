import pandas as pd
from openpyxl import load_workbook

# === Configuration ===
input_file = 'master_list_hrms_august_2025.xlsx'
sheet_name = 'Employee List'
output_file = 'new_emplyee_requests.xlsx'
user_code_column_name = 'User Code'  # Make sure this matches exactly
green_hexes = {'FF00FF00', 'FF92D050', 'FF00B050'} 

# === Step 1: Load data with pandas (for easy filtering/export) ===
df = pd.read_excel(input_file, sheet_name=sheet_name, engine='openpyxl')
df.columns = df.columns.str.strip()  # Clean column names

# === Step 2: Load the same sheet with openpyxl to check formatting ===
wb = load_workbook(input_file)
ws = wb[sheet_name]

# Get header from first row
header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
col_index_map = {col: idx for idx, col in enumerate(header)}

# Find index of 'User Code' column
user_code_col_idx = col_index_map.get(user_code_column_name)
if user_code_col_idx is None:
    raise ValueError(f"Column '{user_code_column_name}' not found in the Excel sheet.")

# === Step 3: Identify matching rows (green cell OR missing/0 user code) ===
matching_row_indices = set()

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):  # min_row=2 skips header
    has_green = any(
        cell.fill.fgColor.rgb and cell.fill.fgColor.type == "rgb" and cell.fill.fgColor.rgb.upper() in green_hexes
        for cell in row
    )

    user_code_cell = row[user_code_col_idx]
    user_code_val = user_code_cell.value
    is_user_code_missing = user_code_val in (None, '', 0, '0')

    if has_green or is_user_code_missing:
        matching_row_indices.add(row_idx)

# === Step 4: Export matching rows (preserve all columns) ===
filtered_df = df.iloc[list(matching_row_indices)]
filtered_df.to_excel(output_file, index=False)

print("✅ Done! Exported {len(filtered_df)} rows to '{output_file}'")
